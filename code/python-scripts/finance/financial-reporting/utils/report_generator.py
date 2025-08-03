#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Report Generator Module for Financial Reporting System

This module provides functions to generate PDF and Excel reports
from financial data and visualizations.

Author: NinjaTech AI
Date: August 2025
"""

import os
from datetime import datetime
import pandas as pd
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

def create_pdf_report(title, data_tables, image_paths, output_path="reports"):
    """
    Generate a PDF report with data tables and images
    
    Args:
        title (str): Report title
        data_tables (dict): Dictionary of table names and DataFrames
        image_paths (dict): Dictionary of image titles and file paths
        output_path (str): Output directory
        
    Returns:
        str: Path to generated PDF file
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_path, exist_ok=True)
    
    # Create filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{title.replace(' ', '_')}_{timestamp}.pdf"
    file_path = os.path.join(output_path, filename)
    
    # Create PDF document
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    styles = getSampleStyleSheet()
    
    # Create custom styles
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,  # Center alignment
        spaceAfter=12
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=10
    )
    
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    # Build content
    content = []
    
    # Add title
    content.append(Paragraph(title, title_style))
    content.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    content.append(Spacer(1, 20))
    
    # Add executive summary
    content.append(Paragraph("Executive Summary", subtitle_style))
    content.append(Paragraph(
        "This report provides a comprehensive analysis of financial performance "
        "based on the latest available data. Key metrics and trends are highlighted "
        "to support decision-making and financial planning.", normal_style))
    content.append(Spacer(1, 20))
    
    # Add data tables
    for table_title, df in data_tables.items():
        content.append(Paragraph(table_title, subtitle_style))
        content.append(Spacer(1, 10))
        
        # Format DataFrame for display
        display_df = df.copy()
        
        # Format numeric columns
        for col in display_df.select_dtypes(include=['float']).columns:
            if 'percentage' in col.lower() or 'pct' in col.lower() or 'change' in col.lower():
                display_df[col] = display_df[col].map(lambda x: f"{x:.2f}%" if pd.notnull(x) else "")
            else:
                display_df[col] = display_df[col].map(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")
        
        # Convert DataFrame to list of lists for ReportLab
        data = [df.columns.tolist()] + display_df.reset_index().values.tolist()
        
        # Create table
        table = Table(data, repeatRows=1)
        
        # Add table style
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        content.append(table)
        content.append(Spacer(1, 20))
    
    # Add images
    for img_title, img_path in image_paths.items():
        content.append(Paragraph(img_title, subtitle_style))
        content.append(Spacer(1, 10))
        
        # Add image
        img = Image(img_path, width=6.5*inch, height=4*inch)
        content.append(img)
        content.append(Spacer(1, 20))
    
    # Add footer with page numbers
    def add_page_number(canvas, doc):
        page_num = canvas.getPageNumber()
        text = f"Page {page_num}"
        canvas.drawRightString(7.5*inch, 0.5*inch, text)
        canvas.drawString(0.5*inch, 0.5*inch, title)
    
    # Build PDF
    doc.build(content, onFirstPage=add_page_number, onLaterPages=add_page_number)
    
    return file_path

def create_excel_report(data_tables, output_path="reports"):
    """
    Generate an Excel report with multiple sheets
    
    Args:
        data_tables (dict): Dictionary of sheet names and DataFrames
        output_path (str): Output directory
        
    Returns:
        str: Path to generated Excel file
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_path, exist_ok=True)
    
    # Create filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Financial_Report_{timestamp}.xlsx"
    file_path = os.path.join(output_path, filename)
    
    # Create Excel writer
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Add summary sheet
        summary_data = {
            'Report Section': list(data_tables.keys()),
            'Records': [len(df) for df in data_tables.values()],
            'Generated': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')] * len(data_tables)
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Write each DataFrame to a different sheet
        for sheet_name, df in data_tables.items():
            # Limit sheet name length to 31 characters (Excel limitation)
            safe_sheet_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_sheet_name)
            
            # Get the worksheet
            worksheet = writer.sheets[safe_sheet_name]
            
            # Auto-adjust columns' width
            for i, col in enumerate(df.columns):
                max_len = max(
                    df[col].astype(str).map(len).max(),  # Length of largest value
                    len(str(col))  # Length of column name
                ) + 2  # Add a little extra space
                
                # Excel column index starts at 1, but openpyxl starts at 0
                # Also, first column is the index which is not in df.columns
                col_idx = i + 2  # +1 for 0-indexing, +1 for index column
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = max_len
            
            # Format header row
            for cell in worksheet["1:1"]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    return file_path

def get_column_letter(col_idx):
    """
    Convert column index to Excel column letter
    
    Args:
        col_idx (int): Column index (1-based)
        
    Returns:
        str: Excel column letter
    """
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

def create_html_report(title, data_tables, image_paths, output_path="reports"):
    """
    Generate an HTML report with data tables and images
    
    Args:
        title (str): Report title
        data_tables (dict): Dictionary of table names and DataFrames
        image_paths (dict): Dictionary of image titles and file paths
        output_path (str): Output directory
        
    Returns:
        str: Path to generated HTML file
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_path, exist_ok=True)
    
    # Create filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{title.replace(' ', '_')}_{timestamp}.html"
    file_path = os.path.join(output_path, filename)
    
    # Start building HTML content
    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>{title}</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                line-height: 1.6;
                margin: 0;
                padding: 20px;
                color: #333;
            }}
            h1 {{
                color: #2c3e50;
                text-align: center;
                padding-bottom: 10px;
                border-bottom: 2px solid #eee;
            }}
            h2 {{
                color: #3498db;
                margin-top: 30px;
            }}
            .timestamp {{
                text-align: center;
                color: #7f8c8d;
                margin-bottom: 30px;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                margin-bottom: 30px;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }}
            th {{
                background-color: #f2f2f2;
                font-weight: bold;
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9;
            }}
            .image-container {{
                margin: 30px 0;
                text-align: center;
            }}
            img {{
                max-width: 100%;
                height: auto;
                border: 1px solid #ddd;
            }}
            .footer {{
                margin-top: 50px;
                text-align: center;
                color: #7f8c8d;
                font-size: 0.9em;
                border-top: 1px solid #eee;
                padding-top: 20px;
            }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <div class="timestamp">Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        
        <h2>Executive Summary</h2>
        <p>
            This report provides a comprehensive analysis of financial performance 
            based on the latest available data. Key metrics and trends are highlighted 
            to support decision-making and financial planning.
        </p>
    """
    
    # Add data tables
    for table_title, df in data_tables.items():
        html_content += f"""
        <h2>{table_title}</h2>
        {df.to_html(classes='data-table', border=0)}
        """
    
    # Add images
    for img_title, img_path in image_paths.items():
        # Get relative path for HTML
        rel_path = os.path.relpath(img_path, output_path)
        html_content += f"""
        <div class="image-container">
            <h2>{img_title}</h2>
            <img src="{rel_path}" alt="{img_title}">
        </div>
        """
    
    # Add footer and close HTML
    html_content += f"""
        <div class="footer">
            &copy; {datetime.now().year} Financial Reporting System
        </div>
    </body>
    </html>
    """
    
    # Write HTML to file
    with open(file_path, 'w') as f:
        f.write(html_content)
    
    return file_path

# Try to import openpyxl for Excel styling (optional)
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    # Define a simple fallback if openpyxl is not available
    def get_column_letter(col_idx):
        """Simple column letter generator"""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result